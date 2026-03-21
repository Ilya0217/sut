"""
Модуль: import_export_service.py
Назначение: Обработка операций импорта и экспорта данных
Автор: Разработчик
Дата создания: 20.03.2026
Требования: FR-09, FR-10, LLR_ImportExportService_ValidateFileStructure_01-02,
             LLR_ImportExportService_GenerateExport_01,
             LLR_ImportExportService_ProcessImport_01
"""

import csv
import io
import json

import openpyxl

from models import Requirement, TraceLink, db, generate_req_id
from services.audit_service import log_action


class ImportExportError(Exception):
    """Базовое исключение для ошибок импорта/экспорта."""

    def __init__(self, code, message, details=None):
        self.code = code
        self.message = message
        self.details = details or []
        super().__init__(message)


REQUIRED_REQ_COLUMNS = [
    "ID системный", "Заголовок", "Текст", "Статус", "Приоритет", "Категория"
]

REQUIRED_LINK_COLUMNS = [
    "Исходное требование", "Целевое требование", "Тип связи"
]


def validate_xlsx_structure(file_stream):
    """
    Назначение: Проверка структуры файла XLSX для импорта
    Id требования: LLR_ImportExportService_ValidateFileStructure_01-02
    Входные данные: file_stream (файловый поток)
    Выходные данные: dict с результатами валидации или исключение
    """
    wb = openpyxl.load_workbook(file_stream, read_only=True)
    sheet_names = wb.sheetnames

    if "Требования" not in sheet_names:
        raise ImportExportError(
            "ERR_MISSING_SHEET",
            "В файле отсутствует лист 'Требования'",
        )
    if "Связи" not in sheet_names:
        raise ImportExportError(
            "ERR_MISSING_SHEET",
            "В файле отсутствует лист 'Связи'",
        )

    req_sheet = wb["Требования"]
    headers = [cell.value for cell in next(req_sheet.iter_rows(max_row=1))]
    missing_columns = []
    for col in REQUIRED_REQ_COLUMNS:
        if col not in headers:
            missing_columns.append(col)
    if missing_columns:
        raise ImportExportError(
            "ERR_MISSING_COLUMN",
            f"Отсутствуют обязательные столбцы: {', '.join(missing_columns)}",
            details=missing_columns,
        )

    req_count = sum(1 for _ in req_sheet.iter_rows(min_row=2)) - 1
    link_sheet = wb["Связи"]
    link_count = sum(1 for _ in link_sheet.iter_rows(min_row=2)) - 1

    wb.close()
    return {
        "is_valid": True,
        "requirements_count": max(req_count, 0),
        "links_count": max(link_count, 0),
        "warnings": [],
    }


def import_from_xlsx(file_stream, project_id, user_id):
    """
    Назначение: Импорт требований и связей из XLSX
    Id требования: LLR_ImportExportService_ProcessImport_01
    Входные данные: file_stream, project_id, user_id
    Выходные данные: dict с результатами импорта
    """
    wb = openpyxl.load_workbook(file_stream)
    req_sheet = wb["Требования"]
    link_sheet = wb["Связи"]

    headers = [cell.value for cell in next(req_sheet.iter_rows(max_row=1))]
    id_map = {}
    imported_reqs = 0
    errors = []

    for row_idx, row in enumerate(req_sheet.iter_rows(min_row=2, values_only=True), 2):
        row_data = dict(zip(headers, row))
        old_id = row_data.get("ID системный", "")

        if not row_data.get("Заголовок"):
            errors.append(f"Строка {row_idx}: отсутствует заголовок")
            continue

        new_system_id = generate_req_id()
        req = Requirement(
            system_id=new_system_id,
            custom_id=row_data.get("ID пользовательский"),
            project_id=project_id,
            title=str(row_data.get("Заголовок", "")),
            text=str(row_data.get("Текст", "")),
            category=str(row_data.get("Категория", "functional")),
            priority=str(row_data.get("Приоритет", "medium")),
            status=str(row_data.get("Статус", "draft")),
            created_by=user_id,
            updated_by=user_id,
        )
        db.session.add(req)
        db.session.flush()

        if old_id:
            id_map[str(old_id)] = req.id
        imported_reqs += 1

    link_headers = [cell.value for cell in next(link_sheet.iter_rows(max_row=1))]
    imported_links = 0

    for row_idx, row in enumerate(link_sheet.iter_rows(min_row=2, values_only=True), 2):
        row_data = dict(zip(link_headers, row))
        source_old = str(row_data.get("Исходное требование", ""))
        target_old = str(row_data.get("Целевое требование", ""))

        source_id = id_map.get(source_old)
        target_id = id_map.get(target_old)

        if source_id is None or target_id is None:
            errors.append(
                f"Связь строка {row_idx}: не найдены требования для связи"
            )
            continue

        link_type = str(row_data.get("Тип связи", "derives_from"))
        if link_type not in TraceLink.VALID_TYPES:
            link_type = "derives_from"

        link = TraceLink(
            source_req_id=source_id,
            target_req_id=target_id,
            link_type=link_type,
            description=str(row_data.get("Описание", "")),
            created_by=user_id,
        )
        db.session.add(link)
        imported_links += 1

    db.session.commit()
    wb.close()

    log_action(
        user_id=user_id,
        action="import_xlsx",
        object_type="project",
        object_id=project_id,
        new_value=f"Импортировано: {imported_reqs} требований, {imported_links} связей",
        context="import",
    )

    return {
        "imported_requirements": imported_reqs,
        "imported_links": imported_links,
        "errors": errors,
    }


def export_to_xlsx(project_id, user_id):
    """
    Назначение: Экспорт требований и связей в XLSX
    Id требования: LLR_ImportExportService_GenerateExport_01
    Входные данные: project_id, user_id
    Выходные данные: BytesIO с файлом XLSX
    """
    requirements = Requirement.query.filter_by(
        project_id=project_id, is_deleted=False
    ).all()
    links = TraceLink.query.join(
        Requirement, TraceLink.source_req_id == Requirement.id
    ).filter(Requirement.project_id == project_id).all()

    wb = openpyxl.Workbook()

    ws_req = wb.active
    ws_req.title = "Требования"
    req_headers = [
        "ID системный", "ID пользовательский", "Заголовок", "Текст",
        "Статус", "Приоритет", "Категория", "Версия",
        "Дата создания", "Дата изменения"
    ]
    ws_req.append(req_headers)

    for req in requirements:
        ws_req.append([
            req.system_id,
            req.custom_id or "",
            req.title,
            req.text,
            req.status,
            req.priority,
            req.category,
            req.version,
            req.created_at.strftime("%Y-%m-%d %H:%M:%S") if req.created_at else "",
            req.updated_at.strftime("%Y-%m-%d %H:%M:%S") if req.updated_at else "",
        ])

    ws_links = wb.create_sheet("Связи")
    link_headers = [
        "Исходное требование", "Целевое требование",
        "Тип связи", "Описание", "Статус"
    ]
    ws_links.append(link_headers)

    for link in links:
        source_sid = link.source_req.system_id if link.source_req else ""
        target_sid = link.target_req.system_id if link.target_req else ""
        ws_links.append([
            source_sid, target_sid,
            link.link_type, link.description, link.status,
        ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    wb.close()

    log_action(
        user_id=user_id,
        action="export_xlsx",
        object_type="project",
        object_id=project_id,
        new_value=f"Экспортировано: {len(requirements)} треб., {len(links)} связей",
    )
    return output


def export_to_csv(project_id, user_id):
    """
    Назначение: Экспорт требований в CSV
    Id требования: Functional.ImportExport.Export
    Входные данные: project_id, user_id
    Выходные данные: строка CSV
    """
    requirements = Requirement.query.filter_by(
        project_id=project_id, is_deleted=False
    ).all()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "ID системный", "ID пользовательский", "Заголовок", "Текст",
        "Статус", "Приоритет", "Категория", "Версия"
    ])
    for req in requirements:
        writer.writerow([
            req.system_id, req.custom_id or "", req.title, req.text,
            req.status, req.priority, req.category, req.version,
        ])

    log_action(
        user_id=user_id,
        action="export_csv",
        object_type="project",
        object_id=project_id,
    )
    return output.getvalue()


def export_to_json(project_id, user_id):
    """
    Назначение: Экспорт требований и связей в JSON
    Id требования: Functional.ImportExport.Export
    Входные данные: project_id, user_id
    Выходные данные: строка JSON
    """
    requirements = Requirement.query.filter_by(
        project_id=project_id, is_deleted=False
    ).all()
    links = TraceLink.query.join(
        Requirement, TraceLink.source_req_id == Requirement.id
    ).filter(Requirement.project_id == project_id).all()

    data = {
        "requirements": [
            {
                "system_id": r.system_id,
                "custom_id": r.custom_id,
                "title": r.title,
                "text": r.text,
                "status": r.status,
                "priority": r.priority,
                "category": r.category,
                "version": r.version,
            }
            for r in requirements
        ],
        "links": [
            {
                "source": lnk.source_req.system_id if lnk.source_req else "",
                "target": lnk.target_req.system_id if lnk.target_req else "",
                "type": lnk.link_type,
                "description": lnk.description,
            }
            for lnk in links
        ],
    }

    log_action(
        user_id=user_id,
        action="export_json",
        object_type="project",
        object_id=project_id,
    )
    return json.dumps(data, ensure_ascii=False, indent=2)


def import_from_csv(file_stream, project_id, user_id):
    """
    Назначение: Импорт требований из CSV
    Id требования: Functional.ImportExport.Import
    Входные данные: file_stream, project_id, user_id
    Выходные данные: dict с результатами
    """
    content = file_stream.read().decode("utf-8")
    reader = csv.DictReader(io.StringIO(content))
    imported = 0
    errors = []

    for row_idx, row in enumerate(reader, 2):
        title = row.get("Заголовок", "").strip()
        if not title:
            errors.append(f"Строка {row_idx}: отсутствует заголовок")
            continue

        req = Requirement(
            system_id=generate_req_id(),
            custom_id=row.get("ID пользовательский", "").strip() or None,
            project_id=project_id,
            title=title,
            text=row.get("Текст", ""),
            category=row.get("Категория", "functional"),
            priority=row.get("Приоритет", "medium"),
            status=row.get("Статус", "draft"),
            created_by=user_id,
            updated_by=user_id,
        )
        db.session.add(req)
        imported += 1

    db.session.commit()

    log_action(
        user_id=user_id,
        action="import_csv",
        object_type="project",
        object_id=project_id,
        new_value=f"Импортировано: {imported} требований",
        context="import",
    )
    return {"imported_requirements": imported, "errors": errors}


def import_from_json(file_stream, project_id, user_id):
    """
    Назначение: Импорт требований и связей из JSON
    Id требования: Functional.ImportExport.Import
    Входные данные: file_stream, project_id, user_id
    Выходные данные: dict с результатами
    """
    content = file_stream.read().decode("utf-8")
    data = json.loads(content)
    id_map = {}
    imported_reqs = 0
    imported_links = 0
    errors = []

    for item in data.get("requirements", []):
        title = item.get("title", "").strip()
        if not title:
            errors.append("Требование без заголовка пропущено")
            continue

        old_id = item.get("system_id", "")
        req = Requirement(
            system_id=generate_req_id(),
            custom_id=item.get("custom_id"),
            project_id=project_id,
            title=title,
            text=item.get("text", ""),
            category=item.get("category", "functional"),
            priority=item.get("priority", "medium"),
            status=item.get("status", "draft"),
            created_by=user_id,
            updated_by=user_id,
        )
        db.session.add(req)
        db.session.flush()
        if old_id:
            id_map[old_id] = req.id
        imported_reqs += 1

    for lnk in data.get("links", []):
        source_id = id_map.get(lnk.get("source"))
        target_id = id_map.get(lnk.get("target"))
        if source_id is None or target_id is None:
            errors.append("Связь пропущена: требования не найдены")
            continue

        link = TraceLink(
            source_req_id=source_id,
            target_req_id=target_id,
            link_type=lnk.get("type", "derives_from"),
            description=lnk.get("description", ""),
            created_by=user_id,
        )
        db.session.add(link)
        imported_links += 1

    db.session.commit()

    log_action(
        user_id=user_id,
        action="import_json",
        object_type="project",
        object_id=project_id,
        new_value=f"Импортировано: {imported_reqs} треб., {imported_links} связей",
        context="import",
    )
    return {
        "imported_requirements": imported_reqs,
        "imported_links": imported_links,
        "errors": errors,
    }
